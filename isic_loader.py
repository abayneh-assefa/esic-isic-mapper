# File: isic_loader.py

import os
import yaml
import openpyxl
from pymongo import MongoClient
from dotenv import load_dotenv
from embedding_utils import get_all_embeddings
from logger import banner, progress, done
from utils import safe_str, subtract_vectors

load_dotenv()

# ─── Load Config ─────────────────────────────────────────────
with open("config.yaml", "r") as f:
    config = yaml.safe_load(f)

mongo_uri = os.getenv("MONGO_URI", config.get("mongo_uri"))
client = MongoClient(mongo_uri)
db = client["industry_mapping"]

# Select collection: ISIC Rev. 4 or Rev. 5
collection_key = os.getenv("ISIC_COLLECTION", "isic")
collection = db[config["collections"].get(collection_key, "isic")]




def load_isic(filepath="data/isic_data_r5.xlsx", models=None, store=True):
    models = models or [
                            # "nomic-embed-text", 
                            "mxbai-embed-large", 
                            # "bge-m3"
        ]
    
    collection.delete_many({})

    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    header = [safe_str(cell.value) for cell in sheet[1]]
    col = {name: idx for idx, name in enumerate(header)}
    total = sheet.max_row - 1
    data = []
    banner(f"📥 Loading ISIC records from {filepath}")

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        desc = safe_str(row[col.get("description")])
        code = safe_str(row[col.get("full_code")])

        section_label = row[col.get("section_label")]
        division_label = row[col.get("division_label")]
        group_label = row[col.get("group_label")]
        
        inclusion = safe_str(row[col.get("explanatory_note_inclusion")])
        exclusion = safe_str(row[col.get("explanatory_note_exclusion")])

        # Semantic input text includes inclusion note
        # positive_text = f"{desc}.{ f' Includes: {inclusion}' if inclusion else ''}"
        positive_text = f"{desc}"
        # positive_text_full = f"{desc}. \n ISIC context: {section_label}, {division_label}, {group_label} \n { f'Includes: {inclusion}' if inclusion else ''}"
        positive_text_full = f"{desc}. \n { f'Includes: {inclusion}' if inclusion else ''}"
        negative_text = exclusion if exclusion else ""

        embeddings = {}

        for model in models:
            pos_embeds = get_all_embeddings(positive_text, model)
            pos_embeds_full  = get_all_embeddings(positive_text_full, model)
            neg_embeds = get_all_embeddings(negative_text, model) if exclusion else {}

            for mode in ["raw", "cosine", "dot"]:
                key = f"embedding_{mode}_{model}"
                pos_vec = pos_embeds_full.get(key, [])
                neg_vec = neg_embeds.get(key, [])
                adjusted = subtract_vectors(pos_vec, neg_vec) if neg_vec else pos_vec
                embeddings[f"{key}_full"] = adjusted

                embeddings[key] = pos_embeds.get(key, [])
                

        record = {
            "sort_order": row[col.get("sort_order")],
            "section": row[col.get("section")],
            "section_label":  section_label,
            "division": row[col.get("division")],
            "division_label":  division_label ,
            "group": row[col.get("group")],
            "group_label":  group_label ,
            "code": row[col.get("code")],
            "level": row[col.get("level")],
            "full_code": code,
            "description": desc,
            "explanatory_note_inclusion": inclusion,
            "explanatory_note_exclusion": exclusion,
            **embeddings
        }
        data.append(record)
        progress(idx, total, prefix="▶ ISIC Embedding")

    if store:
        save_to_mongo(data)

    done(f"Loaded and embedded {len(data)} ISIC records.")
    return data


# ─── Store in MongoDB ────────────────────────────────────────
def save_to_mongo(records):
    collection.insert_many(records)
    done(f"Stored {len(records)} ISIC entries in collection: {collection.name}")

# ─── Script Entry ─────────────────────────────────────────────
if __name__ == "__main__":
    load_isic("data/isic_data_r5.xlsx", None, True)
    
