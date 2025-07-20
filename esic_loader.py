# File: isic_loader.py

import os
import yaml
import openpyxl
from pymongo import MongoClient
from dotenv import load_dotenv
from embedding_utils import get_all_embeddings
from logger import banner, progress, done
from utils import safe_str

load_dotenv()

# ─── Load Config ─────────────────────────────────────────────
with open("config.yaml", "r") as f:
    config = yaml.safe_load(f)

mongo_uri = os.getenv("MONGO_URI", config.get("mongo_uri"))
client = MongoClient(mongo_uri)
db = client["industry_mapping"]

# Select collection: ISIC Rev. 4 or Rev. 5
collection_key = os.getenv("ESIC_COLLECTION", "esic")
collection = db[config["collections"].get(collection_key, "esic")]

# ─── Load ESIC Excel ────────────────────────────────────
def load_esic(filepath="data/esic_data.xlsx", models=None, store=True):
    models = models or [
                        # "nomic-embed-text", 
                        "mxbai-embed-large", 
                        # "bge-m3"
                        ]
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    header = [safe_str(cell.value) for cell in sheet[1]]
    col = {name: idx for idx, name in enumerate(header)}
    total = sheet.max_row - 1
    data = []
    seen = set()
    banner(f"📥 Loading ESIC records from {filepath}")

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        code = safe_str(row[col.get("Code")])
        if not code or code in seen:
            continue
        seen.add(code)

        title = safe_str(row[col.get("Title of category", 2)])

        # Collect embeddings from all models
        embeddings = {}
        for model in models:
            embeddings.update(get_all_embeddings(title, model))

        record = {
            "code": code,
            "title": title,
            "type": safe_str(row[col.get("Type")]),
            "sector": safe_str(row[col.get("Sector")]),
            "division": safe_str(row[col.get("Division")]),
            "major_group": safe_str(row[col.get("Major Group")]),
            "group": safe_str(row[col.get("Group")]),
            "licensing_category": safe_str(row[col.get("Licensing Category")]),
            **embeddings
        }
        data.append(record)
        progress(idx, total, prefix="▶ ESIC Embedding")
    if(store):
     save_to_mongo(data)
   
    done(f"Loaded and embedded {len(data)} ESIC records.")
    return data

# ─── Store in MongoDB ────────────────────────────────────────
def save_to_mongo(records):
    collection.insert_many(records)
    done(f"Stored {len(records)} ESIC entries in collection: {collection.name}")
    

# ─── Script Entry ─────────────────────────────────────────────
if __name__ == "__main__":
    load_esic("data/esic_data.xlsx", None, True)
    

